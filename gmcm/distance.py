#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Date    : 2018-09-15 12:37:51
# @Author  : Bob Liao (codechaser@163.com)
# @Link    : https://github.com/coderchaser

'''
数据来源：高德地图
'''
import requests
import openpyxl
import json
cities = ["乌鲁木齐","北京","哈尔滨","西安","郑州","上海","拉萨","成都","重庆","武汉","广州","昆明"]
city_codes = {}
distance = {}
for i in cities:
	distance[i] = {}


##这里可能会出错
GET_LOCATION = "https://restapi.amap.com/v3/geocode/geo"
GET_DISTANCE= "https://restapi.amap.com/v3/distance"

KEY = "73cb0025ebc81330b0b89d9482316bd3"

LOCATION_PARAMS = {'address':"",'output':"json",'key':KEY,'batch':"true"}

DISTANCE_PARAMETERS = {'origins':"",'output':'json','key':KEY,'destination':"",'type':"0"}


def get_codes():
	first_ten = cities[:10]
	address = '|'.join(first_ten)
	LOCATION_PARAMS['address'] = address
	resp = requests.get(GET_LOCATION,params = LOCATION_PARAMS)
	data = resp.json()['geocodes']
	resp.close()
	last = cities[-2:]
	LOCATION_PARAMS['address'] = '|'.join(last)
	resp = requests.get(GET_LOCATION,params = LOCATION_PARAMS)
	data.extend(resp.json()['geocodes'])
	resp.close()
	codes = [city_dict['location'] for city_dict in data]
	global city_codes
	city_codes = dict(zip(cities,codes))


def get_distance():
	for i in range(len(cities)):
		origin = city_codes[cities[i]]
		DISTANCE_PARAMETERS['origins'] = origin

		for j in range(len(cities)):
			if i == j:
				continue
			destination = city_codes[cities[j]]
			DISTANCE_PARAMETERS['destination'] = destination
			resp = requests.get(GET_DISTANCE,params = DISTANCE_PARAMETERS)
			distance[cities[i]][cities[j]] = round(int(resp.json()['results'][0]['distance'])/1000)
			resp.close()
			

def xlsx_write():
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "表1"
	ws.append(['起点','终点','距离/km'])
	ws.merge_cells(start_row = 2,start_column = 1,end_row = 12,end_column = 1)
	for i in range(len(cities)):
		ws.merge_cells(start_row = 2 + i*11,start_column = 1,end_row = 2 + i*11 + 10,end_column = 1)
	for i in range(len(cities)):
		_ = ws.cell(row = 2+11*i, column = 1, value = cities[i])
		distances = list(distance[cities[i]].items())
		for row in range(2+11*i,2+11*(i+1)):
			_ = ws.cell(row = row,column = 2, value = distances[row - 2 - 11 * i][0])
			_ = ws.cell(row = row,column = 3, value = distances[row - 2 - 11 * i][1])		
		
			
	wb.save('distance.xlsx')


# distance = {'乌鲁木齐': {'北京': 2414, '哈尔滨': 3051, '西安': 2114, '郑州': 2445, '上海': 3271, '拉萨': 1609, '成都': 2067, '重庆': 2308, '武汉': 2768, '广州': 3287, '昆明': 2521}, '北京': {'乌鲁木齐': 2414, '哈尔滨': 1054, '西安': 906, '郑州': 625, '上海': 1068, '拉萨': 2565, '成都': 1526, '重庆': 1460, '武汉': 1054, '广州': 1891, '昆明': 2098}, '哈尔滨': {'乌鲁木齐': 3051, '北京': 1054, '西安': 1961, '郑州': 1644, '上海': 1680, '拉萨': 3556, '成都': 2580, '重庆': 2510, '武汉': 1998, '广州': 2794, '昆明': 3147}, '西安': {'乌鲁木齐': 2114, '北京': 906, '哈尔滨': 1961, '郑州': 432, '上海': 1222, '拉萨': 1757, '成都': 621, '重庆': 578, '武汉': 654, '广州': 1317, '昆明': 1207}, '郑州': {'乌鲁木齐': 2445, '北京': 625, '哈尔滨': 1644, '西安': 432, '上海': 830, '拉萨': 2188, '成都': 1009, '重庆': 881, '武汉': 467, '广州': 1294, '昆明': 1512}, '上海': {'乌鲁木齐': 3271, '北京': 1068, '哈尔滨': 1680, '西安': 1222, '郑州': 830, '拉萨': 2908, '成都': 1663, '重庆': 1444, '武汉': 688, '广州': 1213, '昆明': 1960}, '拉萨': {'乌鲁木齐': 1609, '北京': 2565, '哈尔滨': 3556, '西安': 1757, '郑州': 2188, '上海': 2908, '成都': 1248, '重庆': 1490, '武汉': 2229, '广州': 2318, '昆明': 1272}, '成都': {'乌鲁木齐': 2067, '北京': 1526, '哈尔滨': 2580, '西安': 621, '郑州': 1009, '上海': 1663, '拉萨': 1248, '重庆': 264, '武汉': 981, '广州': 1232, '昆明': 645}, '重庆': {'乌鲁木齐': 2308, '北京': 1460, '哈尔滨': 2510, '西安': 578, '郑州': 881, '上海': 1444, '拉萨': 1490, '成都': 264, '武汉': 756, '广州': 980, '昆明': 638}, '武汉': {'乌鲁木齐': 2768, '北京': 1054, '哈尔滨': 1998, '西安': 654, '郑州': 467, '上海': 688, '拉萨': 2229, '成都': 981, '重庆': 756, '广州': 837, '昆明': 1296}, '广州': {'乌鲁木齐': 3287, '北京': 1891, '哈尔滨': 2794, '西安': 1317, '郑州': 1294, '上海': 1213, '拉萨': 2318, '成都': 1232, '重庆': 980, '武汉': 837, '昆明': 1078}, '昆明': {'乌鲁木齐': 2521, '北京': 2098, '哈尔滨': 3147, '西安': 1207, '郑州': 1512, '上海': 1960, '拉萨': 1272, '成都': 645, '重庆': 638, '武汉': 1296, '广州': 1078}}

get_codes()

get_distance()

xlsx_write()
